# ----encoding='utf-8'-----
#   Chentian's code paradise
import openpyxl
import time
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
'''
⼀、创建
from openpyxl import Workbook
# 实例化
wb = Workbook()
# 获取当前active的sheet
ws = wb.active
print(sheet.title) # 打印sheet表名
sheet.title = "salary luffy" # 改sheet 名

⼆、打开已有⽂件
>>> from openpyxl import load_workbook
>>> wb2 = load_workbook('⽂件名称.xlsx')


2.6.2 写数据
# ⽅式⼀：数据可以直接分配到单元格中(可以输⼊公式)
sheet["C5"] = "Hello ⾦⻆⼤王"
sheet["C7"] = "Hello ⾦⻆⼤王2"
# ⽅式⼆：可以附加⾏，从第⼀列开始附加(从最下⽅空⽩处，最左开始)(可以输⼊多⾏)
sheet.append([1, 2, 3])
# ⽅式三：Python 类型会被⾃动转换
sheet['A3'] = datetime.datetime.now().strftime("%Y-%m-%d")

'''
# wb=Workbook() #内存中创建
#
# ws=wb.active #获取活跃sheet
#
# print(ws)
#
#
# ws.title="陈添的测试文件"
# #按照单元格写入
# ws["B9"]="black girl"
# ws["C9"]="171,48,83"
# #按行写入
# ws.append(["daqin","170","49"])
# ws["A8"]=datetime.datetime.now().strftime("%Y-%m-%d %X")
#
#
#
# wb.save("测试自动生成.xlsx")

wb=load_workbook("测试自动生成.xlsx")

print(wb.sheetnames)


# wb.get_sheet_by_name("陈添的测试文件")
sheet=wb["陈添的测试文件"]
print(sheet["A8"].value)

print(sheet["A1:A14"])

for cell in sheet["A1:A14"]:
    print(cell[0].value)

for row in sheet: # 循环获取表数据
    for cell in row: # 循环获取每个单元格数据
        print(cell.value ,end=",")
    print()
'''
按⾏遍历
for row in sheet: # 循环获取表数据
    for cell in row: # 循环获取每个单元格数据
        print(cell.value, end=",")
    print()
    
按列遍历
# A1, A2, A3这样的顺序
for column in sheet.columns:
    for cell in column:
        print(cell.value,end=",")
    print()
    
遍历指定⾏&列
for row in sheet.iter_rows(min_row=2,max_row=5,max_col=5):
    for cell in row:
        print(cell.value,end=",")
    print()
'''
from openpyxl.styles import Font, colors, Alignment
myfont=Font(name="宋体",size="13",italic=True,color=colors.BLUE)

sheet["A1"].font=myfont

wb.save("自动测试生成")
'''
from openpyxl.styles import Font, colors, Alignment
bold_itatic_24_font = Font(name='等线', size=24, italic=True, color=colors.RED,
bold=True) # 声明样式
sheet['A1'].font = bold_itatic_24_font # 给单元格设置样式
# 设置B1中的数据垂直居中和⽔平居中
sheet['B1'].alignment = Alignment(horizontal='center', vertical='center')
# 第2⾏⾏⾼
sheet.row_dimensions[2].height = 40
# C列列宽
sheet.column_dimensions['C'].width = 30

'''