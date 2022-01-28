# ----encoding='utf-8'-----
#   Chentian's code paradise

from openpyxl import load_workbook
import smtplib
from email.mime.text import MIMEText  # 邮件正文
from email.header import Header  # 邮件头

# 加载excel文件

wb = load_workbook("大唐建设集团-2022年5月工资.xlsx", data_only=True)

sheet = wb["Sheet1"]

smtp_obj = smtplib.SMTP_SSL("smtp.qq.com", 465)
smtp_obj.login("2627601379@qq.com", "lfzrajfynkimecaj")
smtp_obj.set_debuglevel(1)

for row in sheet.iter_rows(min_row=2):
    row_text = "<tr>"
    for cell in row:
        # print(cell.value,end=",")
        row_text += f"<td>{cell.value}</td>"
    row_text += "</tr>"
    name = row[2]
    staffmail = row[1].value
    # print()
    mail_body_context = f'''
     <h3>{name.value},你好：</h3>
     <p>请查收你的工资条</p>
     <table border="1px solid black ">
        {row_text}
     </table>
    '''
    msg = MIMEText("mail_body_context", "html", "utf-8")

    msg["From"] = Header("大唐人事部", "utf-8")
    msg["To"] = Header("大唐员工", "utf-8")
    msg["Subject"] = Header("大唐建设集团2022工资", "utf-8")

    smtp_obj.sendmail("2627601379@qq.com", [staffmail, ], msg.as_string())
